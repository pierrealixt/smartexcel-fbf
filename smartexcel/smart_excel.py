import copy
import xlsxwriter
from openpyxl import load_workbook
import math


SMART_EXCEL_CONFIG = {
    'sheet_names': ['Sheet1', '_data', '_meta'],
    'dump_date_cell_position': 'B1',
    'header_row_cell_position': 'B2'
}


class SmartExcel():
    """
    The SmartExcel class is responsible for two things:
        * Dump data from a definition to a xlsx file.
        * Parse a xlsx and retrieve data from a definition.

    ATTRIBUTES:
    header_row: number of rows used as Header.
    max_row: number of rows where validation is applied.
    meta_worksheet_name: Name of the 'meta' worksheet (used to store timestamp and settings).
    data_worksheet_name = Name of the 'data' worksheet (used to store list options)
    READMODE: If this is set to True, only parsing is available.
    WRITEMODE: If this is set to true, only dumping is available.
    """  # noqa
    header_row = 1
    max_row = 100
    margin_component = 1
    reserved_sheets = ['_data', '_meta']
    READMODE = False
    WRITEMODE = False

    def __init__(
        self,
            definition=None,
            data=None,
            path=None,
            output='template.xlsx'):
        """
        Init a new instance of the SmartExcel class.

        :param definition: A definition of the xlsx (headers, columns, validations).
        :type definition: list

        :param data: An helper class to retrieve data based on the definition.
        :type data: object

        :param path: The path to a xlsx file. Only in READMODE.
        :type path: string

        :param output: The output of a xlsx file. Only in WRITEMODE.
        :type output: str or io.BytesIO()
        """  # noqa

        assert definition and data

        self.sheets = {}
        self.formats = {}
        self.validations = {}
        self.settings = {}
        self.groups = {}

        self.data = data

        if path:
            self.READMODE = True
            self.init_read_mode(definition, path)
        else:
            self.WRITEMODE = True
            self.init_write_mode(definition, output)


    def init_read_mode(self, definition, path):
        """
        Init in READMODE.
        """
        self.build_columns_from_definition(definition)

        self.workbook = load_workbook(path)

        self.meta_config = check_meta_config(self.workbook)

        header = [c['name'] for c in self.columns]

        check_header(
            self.workbook['Sheet1'],
            header,
            self.meta_config['header_row'])

    def init_write_mode(self, definition, output):
        """
        Init in WRITEMODE.
        """

        self.output = output
        self.workbook = xlsxwriter.Workbook(self.output)

        self.add_reserved_sheets()

        self.parse_definition(definition)


    def parse(self):
        """
        Parse a xlsx file according to the definition and returns an list of dict (rows).
        """  # noqa
        assert self.READMODE

        self.parsed_data = []
        for row_index, row in enumerate(self.workbook['Sheet1'].values):
            parsed_row = {}
            if row_index < self.meta_config['header_row']:
                continue
            for col_index, col in enumerate(self.columns):
                if col['index'] > 0 and col['key'].find('--') == -1:
                    new_key = '{key}--{index}'.format(
                        key=col['key'],
                        index=col['index'])
                    col['key'] = new_key

                col_index_base_1 = col_index + 1
                row_index_base_1 = row_index + 1
                value = self.workbook['Sheet1'].cell(
                    column=col_index_base_1,
                    row=row_index_base_1).value

                parsed_row[col['key']] = value

            self.parsed_data.append(parsed_row)
        return self.parsed_data

    def dump(self):
        """
        Dump (render) data into a xlsx file according to the definition.
        """
        assert self.WRITEMODE

        # First, we create the user sheets
        for sheet_key, sheet_data in self.sheets.items():
            if not sheet_data['reserved']:
                sheet_data['fd'] = self.workbook.add_worksheet(sheet_data['name'])

        # Then, we create the reserved sheets
        for sheet_key, sheet_data in self.sheets.items():
            if sheet_data['reserved']:
                sheet_data['fd'] = self.workbook.add_worksheet(sheet_data['name'])
                sheet_data['fd'].protect()
                getattr(
                    self,
                    f"build{sheet_data['name']}"
                )()
                # self.build_meta()
                # self.build_data()

        for sheet_key, sheet_data in self.sheets.items():
            if sheet_data['reserved']:
                continue

            fd_current_sheet = sheet_data['fd']

            # apply settings to each sheet
            for method, value in sheet_data['settings'].items():
                if value is None:
                    getattr(
                        fd_current_sheet,
                        method
                    )()
                else:
                    getattr(
                        fd_current_sheet,
                        method
                    )(value)

            next_available_row = 0

            for component in sheet_data['components']:
                next_available_row += getattr(
                    self,
                    f"render_{component['type']}_component"
                )(fd_current_sheet, component, next_available_row)

        self.workbook.close()

    def render_map_component(self, fd_current_sheet, component, next_available_row):
        """Render a Map component into the current sheet at the next available row.

        :param fd_current_sheet:

        :param component:

        :param next_available_row:
        """
        map_key_format = self.get_component_format(component, 'map_key')
        map_value_format = self.get_component_format(component, 'map_value')

        for index, row in enumerate(component['rows']):
            start_row = next_available_row + index

            try:
                margin_left = component['position']['margin']['left']
            except (KeyError, TypeError):
                margin_left = 0

            start_col = margin_left

            if 'name' in row:
                map_key = row['name']

                fd_current_sheet.write(
                    start_row,
                    start_col,
                    map_key,
                    map_key_format)

                map_value = self.get_value(
                    self.data,
                    f"write_{row['data_func']}",
                    component['payload'][0],
                    {})

                if 'format_func' in row:
                    cell_format = getattr(
                        self.data,
                        f"get_format_for_{row['format_func']}"
                    )(component['payload'][0])

                    map_value_format = self.workbook.add_format(cell_format)

                try:
                    middle = component['position']['middle']
                except (KeyError, TypeError):
                    middle = 1

                start_col += middle

            fd_current_sheet.write(
                start_row,
                start_col,
                map_value,
                map_value_format)

        return len(component['rows']) + self.margin_component

    def render_table_component(self, fd_current_sheet, component, next_available_row):
        """Render a Table component into the current sheet at the next available row.

        :param fd_current_sheet:

        :param component:

        :param next_available_row:
        """
        header_format = self.get_component_format(component, 'header')
        component_cell_format = self.get_component_format(component, 'cell')

        for column in component['columns']:
            self.write_header(
                fd_current_sheet,
                column,
                next_available_row,
                header_format)

            # validations
            self.set_validations(fd_current_sheet, column)

            values = self.get_values_for_column(column, component['payload'])
            self.set_column_width(fd_current_sheet, column, values)

            # format
            cell_format = self.get_column_format(column) or component_cell_format

            for index, value in enumerate(values):
                # TODO: use 0;0 notation instead of A1
                cell_pos = f'{column["letter"]}{index + next_available_row + 1 + self.header_row}'
                fd_current_sheet.write(cell_pos, value, cell_format)

        return len(values) + 1 + self.margin_component

    def write_header(self, sheet, column, next_available_row, header_format):
        col = column["letter"]
        row = self.header_row + next_available_row
        cell_pos = f'{col}{row}'

        # TODO: use 0;0 notation instead of A1
        sheet.write(cell_pos, column['name'], header_format)

    def render_text_component(self, fd_current_sheet, component, next_available_row):
        """Render a Text component into the current sheet at the next available row.

        :param fd_current_sheet:

        :param component:

        :param next_available_row:
        """

        first_row = next_available_row
        first_col = 0

        last_row = first_row + component['size']['height']
        last_col = first_col + (component['size']['width'] - 1)

        range_format = self.get_format(component['format'])

        fd_current_sheet.merge_range(
            first_row,
            first_col,
            last_row,
            last_col,
            component['text'],
            range_format)

        return component['size']['height'] + self.margin_component

    def render_image_component(self, fd_current_sheet, component, next_available_row):
        """Render a Image component into the current sheet at the next available row.

        :param fd_current_sheet: file descriptor of the current sheet
        :type fd_current_sheet: xlsx
        :param component:

        :param next_available_row:
        """
        if 'position' in component and 'x' in component['position']:
            n_row = component['position']['x']
            n_col = component['position']['y']
        else:
            n_row = next_available_row
            n_col = 0

        fd_current_sheet.insert_image(
            n_row, n_col,
            component['image'],
            component['parameters'])

        try:
            component['position']['float']
            return 0
        except KeyError:
            # TODO: this constant should go somewhere else
            height_cell_px = 20
            image_heigt_in_cells = math.ceil(component['size']['height'] / height_cell_px)
            return image_heigt_in_cells + self.margin_component


    def add_reserved_sheets(self):
        """SmartExcel automatically adds two spreadsheets:
        - _data: to store drop-down list values
        - _meta: to store configurations
        """
        for sheet_name in self.reserved_sheets:
            self.sheets[sheet_name] = {
                'reserved': True,
                'name': sheet_name
            }

    def parse_definition(self, definition):
        """Parse a definition.

        :param definition: the spreadhsheet definition
        :type definition: list of dict

        Supported types:
        - sheet
        - format
        """

        for elem in definition:
            try:
                if elem['type'] == 'sheet':
                    self.parse_sheet(elem)
                elif elem['type'] == 'format':
                    self.parse_format(elem)
            except KeyError:
                pass

    def parse_sheet(self, definition, index=0):
        """Parse a sheet definition.

        Attributes:
        - type : 'sheet'
        - key : a string
        - name : either a string or a list:
            => 'name': 'A sheet name'
            => 'name': {
                    'func': 'a_function'
                }
        - components: a list
        - settings: a dict
            => https://xlsxwriter.readthedocs.io/page_setup.html
        """

        # parse sheet name
        try:
            if isinstance(definition['name'], str):
                sheet_name = definition['name']
            else:
                sheet_name = getattr(
                    self.data,
                    f"get_sheet_name_for_{definition['name']['func']}")()
        except KeyError:
            sheet_name = f'Default-{index}'

        if sheet_name in self.reserved_sheets:
            raise ValueError(f'{sheet_name} is a reserved sheet name.')

        # parse sheet key
        try:
            sheet_key = definition['key']
        except KeyError:
            sheet_key = f'{sheet_name}-{index}'

        # parse settings
        if 'settings' in definition:
            settings = definition['settings']
        else:
            settings = {}

        self.sheets[sheet_key] = {
            'name': sheet_name,
            'reserved': False,
            'components': [],
            'settings': settings
        }

        if 'components' in definition:
            kwargs = {
                'sheet_key': sheet_key,
                'settings': settings
            }
            self.parse_components(definition['components'], **kwargs)

    def parse_components(self, components, **kwargs):
        """Parse sheet's components."""

        component_required_attrs = [
            'type',
            'name',
        ]

        for component in components:
            params = copy.deepcopy(kwargs)

            params.update(component)

            validate_attrs(component_required_attrs, params, 'component')

            if component['type'] == 'table':
                self.parse_table(**params)
            elif component['type'] == 'map':
                self.parse_map(**params)
            elif component['type'] == 'text':
                self.parse_text(**params)
            elif component['type'] == 'image':
                self.parse_image(**params)
            else:
                raise ValueError(f"Type `{component['type']}` not supported.")

            if 'recursive' in component:
                self.parse_recursive_components(component, **kwargs)

    def parse_recursive_components(self, parent_comp, **kwargs):
        for index, instance in enumerate(self.data.results[parent_comp['payload']]):

            # compute payload for the current instance.
            payload = self.get_payload(
                func_name=parent_comp['recursive']['payload_func'],
                instance=instance,
                foreign_key=parent_comp['recursive']['foreign_key'])

            self.data.results[parent_comp['recursive']['payload_func']] = payload

            # inherit parent's format
            if 'format' in parent_comp:
                comp_format = parent_comp['format']
            elif 'format' in parent_comp['recursive']:
                comp_format = parent_comp['recursive']['format']
            else:
                comp_format = None

            for child_comp in parent_comp['recursive']['components']:
                # assign parent's `payload` to Map or Table children components.
                # Image or Text components don't need a `payload`.
                if child_comp['type'] in ['map', 'table']:
                    child_comp['payload'] = parent_comp['recursive']['payload_func']

                # assign parent's `format` to children having the same type.
                if child_comp['type'] == parent_comp['type']:
                    child_comp['format'] = comp_format

                # assign the current instance of the payload to the child.
                # similar than `self`.
                child_comp['instance'] = instance

            # compute the child's sheet name
            sheet_name = getattr(
                self.data,
                f"get_sheet_name_for_{parent_comp['recursive']['name']['func']}"
                )(instance)

            definition = {
                'name': sheet_name,
                # children inherit parent's settings
                'settings': kwargs['settings'],
                'components': parent_comp['recursive']['components']
            }

            self.parse_sheet(
                definition=definition,
                index=index)

    def parse_format(self, cell_format):
        """Parse a format.

        A format's definition must be define before a sheet's definition.

        :param cell_format: the format definitiob
        :type cell_format: a dict

        Attributes of `cell_format`:
        - 'type': 'format'
        - 'key': a string (required)
        - 'format': a dict (required):
            => https://xlsxwriter.readthedocs.io/format.html
        """
        required_attrs = [
            'key',
            'format'
        ]

        validate_attrs(required_attrs, cell_format, 'format')

        self.formats[cell_format['key']] = self.workbook.add_format(
            cell_format['format'])

        if 'num_format' in cell_format:
            # It controls whether a number is displayed
            # as an integer, a floating point number, a date,
            # a currency value or some other user defined format.
            self.formats[cell_format['key']].set_num_format(
                cell_format['num_format'])

    def parse_map(self, **kwargs):
        """Parse a Map component (as the data structuce).

        :param kwargs: the map definition
        :type kwargs: a dict

        Attributes:
        - 'type': 'table' (required)
        - 'name': a string (required)
        - 'rows': a list of dict (required)
        - 'payload': a string (required)
        - 'position': a dict
            => {
                'x': 0,
                'y': 0
            }
        - 'format': a dict
            => {
                'map_key',
                'map_value'
            }
        """

        map_required_attrs = [
            'rows',
            'payload'
        ]
        validate_attrs(map_required_attrs, kwargs, 'map component')

        parsed_rows = []
        for row in kwargs['rows']:
            tmp = copy.deepcopy(row)

            tmp.update({
                'letter': 'A'
            })

            parsed_rows.append(tmp)

        if 'format' in kwargs:
            map_format = kwargs['format']
        else:
            map_format = None

        sheet_key = kwargs['sheet_key']

        if 'position' in kwargs:
            position = kwargs['position']
        else:
            position = None

        self.sheets[sheet_key]['components'].append({
            'type': 'map',
            'payload': self.data.results[kwargs['payload']],
            'name': kwargs['name'],
            'rows': parsed_rows,
            'position': position,
            'format': map_format
        })

    def parse_table(self, **kwargs):
        """Parse a Table component.

        Attributes:
        - 'type': 'table' (required)
        - 'name': a string (required)
        - 'columns': a list of dict (required)
        - 'payload': a string (required)
        - 'group_name'
        - 'repeat': a string or a dict
        - 'format': a dict
            => {
                'header': a string
                'cell': a string
            }
        """

        table_required_attrs = [
            'columns',
            'payload'
        ]

        validate_attrs(table_required_attrs, kwargs, 'table component')

        # parse 'group_name'
        if 'group_name' in kwargs:
            group_name = kwargs['group_name']
        else:
            group_name = None

        # parse 'repeat'
        if 'repeat_func' in kwargs:
            repeat = getattr(
                self.data,
                'write_{key}'.format(
                    key=kwargs['repeat_func']))()
        elif 'repeat' in kwargs:
            repeat = kwargs['repeat']
        else:
            repeat = 1

        # parse 'columns'
        parsed_columns = self.parse_columns(
            kwargs['columns'],
            repeat)

        sheet_key = kwargs['sheet_key']

        # parse 'payload'
        if kwargs['payload'] in self.data.results:
            payload = self.data.results[kwargs['payload']]
        else:
            raise ValueError(f"{kwargs['payload']} not present in {self.data} `results` list.")

        # parse format
        if 'format' in kwargs:
            table_format = kwargs['format']
        else:
            table_format = None

        self.sheets[sheet_key]['components'].append({
            'type': 'table',
            'payload': payload,
            'name': kwargs['name'],
            'columns': parsed_columns,
            'format': table_format
        })


    def parse_text(self, **kwargs):
        """Parse a Text component.
        Attributes:
        - 'type': 'table' (required)
        - 'name': a string (required)
        - 'size': a dict (required)
        - 'text_func': a string (required)
        - 'format': a string
        """

        required_attrs = [
            'name',
            'text_func',
            'size',
        ]

        validate_attrs(required_attrs, kwargs, 'text component')
        validate_size(kwargs)

        sheet_key = kwargs['sheet_key']

        if 'format' in kwargs:
            text_format = kwargs['format']
        else:
            text_format = None

        # compute `text` from `text_func`
        if 'instance' in kwargs:
            # `instance` is available for a child's component
            text = getattr(
                self.data,
                f"get_text_for_{kwargs['text_func']}"
            )(kwargs['instance'])
        else:
            text = getattr(
                self.data,
                f"get_text_for_{kwargs['text_func']}"
            )()

        self.sheets[sheet_key]['components'].append({
            'type': 'text',
            'text': text,
            'size': kwargs['size'],
            'format': text_format
        })

    def parse_image(self, **kwargs):
        """Parse a Image component.
        Attributes:
        - 'type': 'table' (required)
        - 'name': a string (required)
        - 'size': a dict (required)
        - 'image_func': a string (required)
        - 'position': a dict
            => {
                'x': int,
                'y': int,
                'float': boolean
            }
        - 'parameters' : a dict
            => https://xlsxwriter.readthedocs.io/worksheet.html#worksheet-insert-image
        """
        required_attrs = [
            'name',
            'image_func',
            'size',
        ]

        validate_attrs(required_attrs, kwargs, 'image component')
        validate_size(kwargs)

        sheet_key = kwargs['sheet_key']

        if 'instance' in kwargs:
            # `instance` is available for a child's component
            image = getattr(
                self.data,
                f"get_image_{kwargs['image_func']}"
            )(kwargs['instance'], kwargs['size'])
        else:
            image = getattr(
                self.data,
                f"get_image_{kwargs['image_func']}"
            )(kwargs['size'])

        if 'parameters' in kwargs:
            parameters = kwargs['parameters']
        else:
            parameters = {}

        if 'position' in kwargs:
            position = kwargs['position']
        else:
            position = {}

        self.sheets[sheet_key]['components'].append({
            'type': 'image',
            'image': image,
            'size': kwargs['size'],
            'parameters': parameters,
            'position': position
        })


    def parse_columns(self, columns, repeat):
        """Parse columns of a Table component.

        :param columns: the columns definition
        :type columns: list

        :param repeat:
        :type repeat: integer

        Attributes:
        - 'name': a string or a list (required):
            => 'name': 'A sheet name'
            => 'name': {
                    'func': a string
                }
        - 'data_func': a string (required)
        - 'width': a integer
        - 'format': a string
        """

        required_attrs = [
            'name',
            'data_func'
        ]

        parsed_columns = []
        for index in range(0, repeat):
            for column in columns:
                validate_attrs(required_attrs, column, 'column')

                tmp_col = copy.deepcopy(column)

                if isinstance(tmp_col['name'], dict):
                    tmp_col['name'] = self.get_value(
                        klass=self.data,
                        func='write_{key}'.format(key=tmp_col['name']['func']),
                        obj=None,
                        kwargs={'index': index})

                if repeat > 1:
                    name = f'{tmp_col["name"]} - {index + 1}'
                else:
                    name = tmp_col['name']

                tmp_col.update({
                    'name': name,
                    'letter': next_letter(len(parsed_columns)),
                    'index': index
                })

                parsed_columns.append(tmp_col)
            return parsed_columns


    def set_list_source_func(self, sheet, cell_range, column):
        if 'validations' in column and 'list_source_func' in column['validations']:
            sheet.data_validation(cell_range, {
                'validate': 'list',
                'source': f'={self.validations[column["key"]]["meta_source"]}'
            })

    def column_cell_range(self, column):
        return '{start_letter}{start_pos}:{end_letter}{end_pos}'.format(
            start_letter=column["letter"],
            start_pos=self.header_row + 1,
            end_letter=column["letter"],
            end_pos=self.max_row
        )

    def set_validations(self, sheet, column):
        if column['data_func'] in self.validations:
            cell_range = self.column_cell_range(column)

            self.set_list_source_func(sheet, cell_range, column)
            self.set_excel_validations(sheet, cell_range, column)

    def set_excel_validations(self, sheet, cell_range, column):
        if 'validations' in column and 'excel' in column['validations']:
            sheet.data_validation(
                cell_range,
                column['validations']['excel'])

    def build_meta(self):
        """Populate the reserved sheet name `_meta`.
        """
        from datetime import datetime
        now = datetime.now().strftime('%Y-%m-%d')
        self.sheets['_meta']['fd'].write_row('A1', ['dump_date', now])
        self.sheets['_meta']['fd'].write_row('A2', ['header_rows', self.header_row])

    def build_data(self):
        """Populate the reserved sheet name `_data`.
        """
        for sheet_key, sheet_data in self.sheets.items():
            if sheet_data['reserved']:
                continue
            for component in sheet_data['components']:
                if 'rows' in component:
                    pass
                if 'columns' in component:
                    for column in component['columns']:
                        if column['data_func'] not in self.validations:
                            tmp = {
                                'row': len(self.validations) + 1
                            }
                            if 'validations' in column:
                                tmp.update(column['validations'])

                                if 'list_source_func' in column['validations']\
                                    and column['data_func'] not in self.validations:
                                    list_source = getattr(
                                        self.data,
                                        column['validations']['list_source_func'])()

                                    tmp.update({
                                        'meta_source': f'={self.data_worksheet_name}!$A${tmp["row"]}:${next_letter(len(list_source) - 1)}${tmp["row"]}'  # noqa
                                    })

                                    self.sheets['_data']['fd'].write_row(
                                        f"A{tmp['row']}",
                                        list_source)

                                self.validations[column['data_func']] = tmp

    def get_format(self, format_name):
        if format_name in self.formats:
            return self.formats[format_name]
        else:
            None

    def get_column_format(self, column):
        if 'format' in column:
            return self.formats[column['format']]
        else:
            return None

    def get_component_format(self, component, format_type):
        if 'format' in component and component['format']:
            if format_type in component['format']:
                return self.get_format(component['format'][format_type])


    def build_top_header(self):
        if self.groups:
            self.header_row += 1
            for group_name, position in self.groups.items():
                if group_name is None:
                    continue

                if position['start'] == position['end']:
                    self.main_ws.set_column(
                        f"{position['start']}:{position['start']}",
                        len(group_name))

                    self.main_ws.write(
                        f"{position['start']}1",
                        group_name,
                        self.get_format('top_header'))

                else:
                    cell_range = f"{position['start']}1:{position['end']}1"
                    self.main_ws.merge_range(
                        cell_range,
                        group_name,
                        self.get_format('top_header'))

        self.main_ws.freeze_panes(1, 0)
        self.main_ws.freeze_panes(2, 0)

    def set_column_width(self, fd_current_sheet, column, values):
        if 'width' in column:
            width = column['width']
        else:
            try:
                width = len(max(values, key=len))
                if width < 10:
                    width = 10
            except (TypeError, ValueError):
                width = 10

        fd_current_sheet.set_column(
            f"{column['letter']}:{column['letter']}",
            width)

    def get_meta(self, klass, func, meta, kwargs):
        if func not in dir(klass):
            raise Exception(f'method \'{func}\' not present in {klass} class')  # noqa
        try:
            meta = getattr(klass, func)(meta, kwargs)
        except IndexError:
            meta = None

        return meta

    def get_value(self, klass, func, obj, kwargs):
        return self.get_meta(klass, func, obj, kwargs)

    def get_values_for_column(self, column, payload):
        return [
            self.get_value(
                self.data,
                'write_{key}'.format(key=column['data_func']),
                obj,
                {'index': index})
            for index, obj in enumerate(payload)  # self.data.results
        ]

    def get_payload(self, func_name, instance, foreign_key):
        """Calling the payload method `func_name` on the DataModel class.

        The DataModel class is expected to define this function:
        e.g: `func_name` = 'detail'
        ```
        def get_payload_detail(self, instance, foreign_key):
            return 42
        ```

        :param func_name: identifier of a function
        :type func_name: str

        :param instance:
        :type instance: dict | tuple | list

        :param foreign_key: a key to link parent instance and its children.
        :type foreign_key: str
        """

        func_name = f'get_payload_{func_name}'
        payload = getattr(
            self.data,
            func_name)(
                instance=instance,
                foreign_key=foreign_key
            )
        return payload


def check_sheet_names(sheet_names):
    if sheet_names != SMART_EXCEL_CONFIG['sheet_names']:
        raise Exception("'Sheet1', '_meta', '_data' sheets must be present.")


def check_dump_date(meta_ws):
    dump_date = meta_ws[SMART_EXCEL_CONFIG['dump_date_cell_position']]
    if dump_date is None:
        raise Exception("A dump date must be present.")
    return dump_date.value


def check_header_row(meta_ws):
    header_row = meta_ws[SMART_EXCEL_CONFIG['header_row_cell_position']]
    if header_row is None:
        raise Exception("config header_row must be present.")
    return header_row.value


def check_header(sheet, definition, header_row):
    header_row_base_0 = header_row - 1
    # sheet.values is a generator, so an iterator. I should use itertools
    for index, row in enumerate(sheet.values):
        if index == header_row_base_0:
            if tuple(definition) != row:
                raise Exception("Header definitions do not match.")


def check_meta_config(wb):
    check_sheet_names(wb.sheetnames)

    return {
        'dump_date': check_dump_date(wb['_meta']),
        'header_row': check_header_row(wb['_meta'])
    }



A, Z = 65, 90
TOTAL = 26


def next_letter(length):
    """
    Get the next excel column available.
    e.g: 'A', 'AA', 'BA'

    length: of the columns list
    returns
    """
    pos = int(length / TOTAL)

    if length >= TOTAL:
        next_length = length - (TOTAL * pos)
        return f'{next_letter(pos - 1)}{next_letter(next_length)}'
    else:
        char = length + A
    return chr(char)

def validate_attrs(required_attrs, element, element_type):
    for attr in required_attrs:
        if attr not in element:
            raise ValueError(f'{attr} is required in a {element_type} definition.')


def validate_type(element, attr, required_type):
    if not isinstance(element[attr], required_type):
        raise ValueError(f'{attr} must be a {required_type}')


def validate_position(element):
    attr = 'position'
    keys = ['x', 'y']

    validate_type(element, attr, dict)
    validate_attrs(keys, element[attr], f'component {attr}')
    for key in keys:
        validate_type(element[attr], key, int)

    return True


def validate_size(element):
    attr = 'size'
    keys = ['width', 'height']

    validate_type(element, attr, dict)
    validate_attrs(keys, element[attr], f'component {attr}')
    for key in keys:
        validate_type(element[attr], key, int)

    return True
